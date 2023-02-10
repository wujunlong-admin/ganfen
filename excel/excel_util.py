from openpyxl import load_workbook
from excel import case


class ExcelTuil:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)

    # 读取Excel中测试数据
    def read_data(self, sheet_name):
        sheet = self.wb[sheet_name]
        cases = []
        for i in range(1, sheet.max_row + 1):
            test_data = case.Case()
            test_data.id = sheet.cell(i, 1).value
            test_data.name = sheet.cell(i, 2).value
            test_data.age = sheet.cell(i, 3).value
            test_data.password = sheet.cell(i, 4).value
            cases.append(test_data)
        return cases

    # 测试结果写入Excel中
    def write_data(self, sheet_name, row, col, value):
        sheet = self.wb[sheet_name]
        sheet.cell(row, col, value)
        self.wb.save(self.file_name)
